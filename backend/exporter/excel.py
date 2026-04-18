"""Excel exporter with 8 sheets."""
from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HEADERS = [
    ("№", 6),
    ("Компания", 30),
    ("Сайт", 22),
    ("ИНН", 14),
    ("КПП", 12),
    ("Общий email", 26),
    ("Общий телефон", 20),
    ("ФИО (полностью)", 30),
    ("Фамилия", 18),
    ("Имя", 14),
    ("Отчество", 18),
    ("Пол", 6),
    ("Должность (как на сайте)", 32),
    ("Должность (норм.)", 28),
    ("Категория отрасли", 22),
    ("Метод нормализации", 14),
    ("Личный email", 26),
    ("Личный телефон", 20),
    ("Соцсети", 30),
    ("URL источника", 38),
    ("Язык", 8),
    ("Дата", 12),
    ("Статус", 10),
    ("Комментарий", 24),
]

SHEET_NAMES = [
    "Генеральные директора",
    "Финансовые директора",
    "Главные бухгалтеры",
    "Главные инженеры",
    "Остальные",
    "Все контакты",
    "Сводка",
    "Отчёт качества",
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_PARTIAL_FILL = PatternFill("solid", fgColor="FEF3C7")
_ERROR_FILL = PatternFill("solid", fgColor="FEE2E2")


def _setup_sheet(ws):
    for col_idx, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _ALIGN
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 24


def _contact_row(c: Dict, n: int) -> list:
    socials = c.get("social_links") or []
    if isinstance(socials, list):
        socials = "\n".join(socials)
    return [
        n,
        c.get("company_name") or "",
        c.get("domain") or "",
        c.get("inn") or "",
        c.get("kpp") or "",
        c.get("company_email") or "",
        c.get("company_phone") or "",
        c.get("full_name") or "",
        c.get("last_name") or "",
        c.get("first_name") or "",
        c.get("patronymic") or "",
        c.get("gender") or "",
        c.get("position_raw") or "",
        c.get("position_canonical") or "",
        c.get("role_category") or "",
        c.get("norm_method") or "",
        c.get("person_email") or "",
        c.get("person_phone") or "",
        socials,
        c.get("page_url") or "",
        c.get("language") or "",
        c.get("scan_date") or "",
        c.get("status") or "ok",
        c.get("comment") or "",
    ]


def _write_contacts(ws, contacts: List[Dict]):
    _setup_sheet(ws)
    for idx, c in enumerate(contacts, start=1):
        row = _contact_row(c, idx)
        for col_idx, v in enumerate(row, start=1):
            cell = ws.cell(row=idx + 1, column=col_idx, value=v)
            cell.alignment = _ALIGN
            cell.border = _BORDER
            if c.get("status") == "error":
                cell.fill = _ERROR_FILL
            elif c.get("status") == "partial":
                cell.fill = _PARTIAL_FILL


def _write_summary(ws, contacts: List[Dict], per_sheet: Dict[str, List[Dict]], task_meta: Dict):
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    rows = [
        ("Задача", task_meta.get("task_id", "")),
        ("Создано", task_meta.get("created_at", "")),
        ("Статус", task_meta.get("status", "")),
        ("Всего URL", task_meta.get("total_urls", 0)),
        ("Обработано URL", task_meta.get("processed_urls", 0)),
        ("", ""),
        ("Всего записей", len(contacts)),
        ("Уникальных доменов", len({c.get("domain") for c in contacts if c.get("domain")})),
        ("С ФИО", sum(1 for c in contacts if c.get("full_name"))),
        ("С личным email", sum(1 for c in contacts if c.get("person_email"))),
        ("С личным телефоном", sum(1 for c in contacts if c.get("person_phone"))),
        ("С ИНН", sum(1 for c in contacts if c.get("inn"))),
        ("", ""),
    ]
    for name in ["Генеральные директора", "Финансовые директора", "Главные бухгалтеры",
                 "Главные инженеры", "Остальные"]:
        rows.append((f"Лист: {name}", len(per_sheet.get(name, []))))
    rows.append(("", ""))
    methods = Counter(c.get("norm_method") or "unknown" for c in contacts)
    for m in ("exact", "morph", "fuzzy", "fallback", "empty"):
        rows.append((f"Норм. {m}", methods.get(m, 0)))

    for i, (k, v) in enumerate(rows, start=1):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=v)
        if k and not k.startswith(" "):
            c1.font = Font(bold=k in {"Задача", "Всего записей", "Методы нормализации:"} or "Лист:" in k)
        c1.alignment = _ALIGN
        c2.alignment = _ALIGN


def _write_quality(ws, contacts: List[Dict], errors: List[Dict]):
    ws.append(["Компания", "Сайт", "Сырая должность", "Нормализованная", "Метод", "Комментарий"])
    for col_idx, (_, w) in enumerate(zip(range(6), [30, 22, 34, 30, 14, 30])):
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = w
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    # Sort by method: fallback first
    order = {"fallback": 0, "fuzzy": 1, "morph": 2, "exact": 3, "empty": -1}
    rows = sorted(contacts, key=lambda c: order.get(c.get("norm_method"), 99))
    for c in rows:
        ws.append([
            c.get("company_name") or "",
            c.get("domain") or "",
            c.get("position_raw") or "",
            c.get("position_canonical") or "",
            c.get("norm_method") or "",
            c.get("comment") or "",
        ])
    # Errors section
    if errors:
        ws.append([])
        ws.append(["Ошибки сайтов:"])
        ws.append(["URL", "Ошибка", "Сообщение"])
        for e in errors:
            ws.append([e.get("url", ""), e.get("error_code", ""), e.get("error_message", "")])


def _bucket_by_sheet(contacts: List[Dict]) -> Dict[str, List[Dict]]:
    out = {name: [] for name in SHEET_NAMES if name not in {"Все контакты", "Сводка", "Отчёт качества"}}
    for c in contacts:
        s = c.get("sheet_name") or "Остальные"
        if s not in out:
            s = "Остальные"
        out[s].append(c)
    return out


def generate_excel(
    contacts: List[Dict],
    output_path: str,
    task_meta: Dict,
    errors: List[Dict] | None = None,
) -> str:
    """Generate Excel workbook with 8 sheets. Returns output path."""
    errors = errors or []
    wb = Workbook()
    # Remove default
    default = wb.active
    wb.remove(default)

    # Enrich contacts with scan_date if missing
    now = datetime.utcnow().strftime("%Y-%m-%d")
    for c in contacts:
        c.setdefault("scan_date", now)

    per_sheet = _bucket_by_sheet(contacts)

    # Role sheets + "Остальные"
    for name in ["Генеральные директора", "Финансовые директора", "Главные бухгалтеры",
                 "Главные инженеры", "Остальные"]:
        ws = wb.create_sheet(name)
        _write_contacts(ws, per_sheet.get(name, []))

    # "Все контакты"
    ws_all = wb.create_sheet("Все контакты")
    _write_contacts(ws_all, contacts)

    # Сводка
    ws_sum = wb.create_sheet("Сводка")
    _write_summary(ws_sum, contacts, per_sheet, task_meta)

    # Отчёт качества
    ws_q = wb.create_sheet("Отчёт качества")
    _write_quality(ws_q, contacts, errors)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
